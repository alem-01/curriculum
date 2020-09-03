import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')

print(wb.sheetnames)

sheet = wb['Instructions']

print(sheet.title)

vl = sheet['A10']

print(vl.row, vl.column, vl.value)


for i in range(5, 25, 3):
    print(i, sheet.cell(row = i, column = 5).value)

print(sheet.max_row, sheet.max_column)

print(get_column_letter(sheet.max_column))